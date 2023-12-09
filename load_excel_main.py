from openpyxl import load_workbook

workbook = load_workbook(filename='my_table2.xlsx')
print(workbook.sheetnames)

sheet = workbook.worksheets[1]
print(sheet)
# sheet = workbook.active
print('test row column ---', sheet.cell(row=1, column=2).value)

print(sheet['A1'].value)

print('----TEST FOR LOOP ---')

for index in range(1, sheet.max_column + 1):
    cell_obj = sheet.cell(row=1, column=index)
    print(cell_obj.value)

print(sheet[1])

print('---iter_rows----')

for row in sheet.iter_rows(max_row=2, min_row=1, min_col=1, max_col=4, values_only=True):
    print(row)