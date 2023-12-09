from openpyxl import Workbook


name = 'my_table2.xlsx'

wb = Workbook()

sheet = wb.active
sheet['A1'] = 'Hello'
sheet['B1'] = 'strange'
sheet['C1'] = 'World'
sheet['D1'] = "I'm Max"

sheet = wb.active
sheet['A2'] = 'Hello'
sheet['B2'] = 'wonderful'
sheet['C2'] = 'Planet'
sheet['D2'] = "I'm not Max"

sheet['B3'] = 'Do you know something interesting?'

sheet.title = 'Dci Student Sheet'

ws1 = wb.create_sheet('Strange_sheet')
ws2 = wb.create_sheet('Sad_sheet', 0)


wb.save(filename=name)